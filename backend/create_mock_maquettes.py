import os
import shutil
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output_maquettes")
TEST_DIR = os.path.join(OUTPUT_DIR, "test_mocks")

if not os.path.exists(TEST_DIR):
    os.makedirs(TEST_DIR)

original_file = os.path.join(OUTPUT_DIR, "MAQUETTE_A_AFFECTATION_MSD.xlsx")

if not os.path.exists(original_file):
    print(f"File not found: {original_file}")
    exit(1)

# 1. Base copy
mock1_path = os.path.join(TEST_DIR, "MOCK_1_BASE.xlsx")
shutil.copy(original_file, mock1_path)
print(f"Created {mock1_path}")

# 2. Add some prof assignments
mock2_path = os.path.join(TEST_DIR, "MOCK_2_ASSIGNED_PROFS.xlsx")
wb = load_workbook(original_file)
ws = wb.active
# Usually row 5 is where data starts (header is row 4, title 1)
# Let's find first row with data
for row in range(5, ws.max_row + 1):
    val = ws.cell(row=row, column=3).value # Module Name
    if val:
        # Assign profs
        ws.cell(row=row, column=5).value = "M. ALAOUI, MME. KADIRI"
        ws.cell(row=row+1, column=5).value = "M. TAZI"
        break
wb.save(mock2_path)
print(f"Created {mock2_path}")

# 3. Add a completely new course
mock3_path = os.path.join(TEST_DIR, "MOCK_3_ADDED_COURSE.xlsx")
wb = load_workbook(original_file)
ws = wb.active
new_row_idx = ws.max_row + 1
ws.cell(row=new_row_idx, column=1).value = f"NEW_PART_{new_row_idx}"
ws.cell(row=new_row_idx, column=2).value = "MSD S4"
ws.cell(row=new_row_idx, column=3).value = "Intelligence Artificielle Avancée (CM)"
ws.cell(row=new_row_idx, column=4).value = "24H"
ws.cell(row=new_row_idx, column=5).value = "M. EL FASSY"
wb.save(mock3_path)
print(f"Created {mock3_path}")

# 4. Delete some courses
mock4_path = os.path.join(TEST_DIR, "MOCK_4_DELETED_COURSE.xlsx")
wb = load_workbook(original_file)
ws = wb.active
if ws.max_row > 10:
    ws.delete_rows(6, 3) # Delete 3 rows starting from row 6
wb.save(mock4_path)
print(f"Created {mock4_path}")

# 5. Mixed (Add TD + Assign Profs)
mock5_path = os.path.join(TEST_DIR, "MOCK_5_MIXED.xlsx")
wb = load_workbook(original_file)
ws = wb.active
new_row_idx = ws.max_row + 1
ws.cell(row=new_row_idx, column=1).value = f"NEW_TD_{new_row_idx}"
ws.cell(row=new_row_idx, column=2).value = "MSD S2"
ws.cell(row=new_row_idx, column=3).value = "Programmation Python (TD)"
ws.cell(row=new_row_idx, column=4).value = "12H"
ws.cell(row=new_row_idx, column=5).value = "M. CHRAIBI"
# also assign prof
ws.cell(row=5, column=5).value = "PR. OUAZANI"
wb.save(mock5_path)
print(f"Created {mock5_path}")

print("All mocks generated successfully in output_maquettes/test_mocks/")
