import pandas as pd
from sqlalchemy import create_engine, text
import os
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Configuration base de données
DATABASE_URL = "postgresql://user_pfe:password_pfe@localhost:5432/fstm_timetable"
engine = create_engine(DATABASE_URL)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output_maquettes")

if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

def finalize_excel(filename, title_text):
    wb = load_workbook(filename)
    ws = wb.active
    
    navy_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin', color="000000"), right=Side(style='thin', color="000000"), 
                    top=Side(style='thin', color="000000"), bottom=Side(style='thin', color="000000"))

    # TITRE
    ws.merge_cells('A1:E1')
    ws['A1'] = title_text.upper()
    ws['A1'].font = Font(size=18, bold=True, color="1A365D")
    ws['A1'].alignment = Alignment(horizontal="center")

    # EN-TÊTES (LIGNE 4) - 5 COLONNES UNIQUEMENT
    headers = ["ID", "SECTION", "NOM DU MODULE", "VOLUME H.", "ENSEIGNANT"]
    for i, title in enumerate(headers):
        cell = ws.cell(row=4, column=i+1)
        cell.value = title
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # MASQUAGE ID (COLONNE A)
    ws.column_dimensions['A'].width = 0
    ws.column_dimensions['A'].hidden = True

    # LARGEURS
    ws.column_dimensions['B'].width = 25 # SECTION
    ws.column_dimensions['C'].width = 50 # MODULE
    ws.column_dimensions['D'].width = 15 # VH
    ws.column_dimensions['E'].width = 40 # ENSEIGNANT

    # Bordures pour les données
    for row in ws.iter_rows(min_row=5, max_col=5):
        for cell in row:
            cell.border = border

    wb.save(filename)

def generate_maquette_a(f_name):
    # GÉNÉRATION POUR LA FILIÈRE SPÉCIFIÉE
    print(f"⏳ Génération de la Maquette A pour {f_name}...")
    
    query = text("""
        SELECT DISTINCT
            mp.id::text || '_' || s.id::text || '_' || COALESCE(tdg.id::text, '') as internal_id, 
            s.name, 
            CASE 
                WHEN mp.type IN ('TD', 'TP') AND tdg.id IS NOT NULL THEN m.name || ' (' || mp.type || ') - ' || tdg.name
                ELSE m.name || ' (' || mp.type || ')'
            END as module_col,
            mp.weekly_hours, 
            COALESCE(
                (SELECT string_agg(DISTINCT t.name, ', ') 
                 FROM public.assignments a 
                 JOIN public.teachers t ON a.teacher_id = t.id 
                 LEFT JOIN public.assignment_tdgroups atd ON a.id = atd.assignment_id
                 WHERE a.module_part_id = mp.id 
                   AND (
                       (mp.type = 'CM' AND (a.section_id = s.id OR atd.tdgroup_id IN (SELECT id FROM public.td_groups WHERE section_id = s.id))) OR 
                       (mp.type IN ('TD', 'TP') AND atd.tdgroup_id = tdg.id)
                   )
                ), ''
            ) as existing_teacher,
            s.name as s1, 
            m.name as m1,
            mp.type as t1,
            COALESCE(tdg.name, '') as td_name
        FROM public.sections s
        JOIN public.section_groupes sg ON s.id = sg.section_id
        JOIN public.groupe_filieres gf ON sg.groupe_id = gf.id
        CROSS JOIN LATERAL (
            SELECT sg2.section_id FROM public.section_groupes sg2 
            JOIN public.groupe_filieres gf2 ON sg2.groupe_id = gf2.id 
            WHERE gf2.filiere_id = (SELECT id FROM filieres WHERE name = :filiere_name)
        ) ts
        JOIN public.groupe_module_groupes gmg ON gf.id = gmg.groupe_id
        JOIN public.groupe_modules gm ON gmg.groupe_module_id = gm.id
        JOIN public.modules m ON gm.module_id = m.id
        JOIN public.module_parts mp ON m.id = mp.module_id
        LEFT JOIN public.td_groups tdg ON tdg.section_id = s.id AND mp.type IN ('TD', 'TP')
        WHERE s.id = ts.section_id
        ORDER BY s1, m1, t1, td_name
    """)
    
    df = pd.read_sql(query, engine, params={"filiere_name": f_name})
    if df.empty: 
        print(f"❌ Aucun module trouvé pour {f_name}.")
        return

    # Préparation du format Excel
    df = df.iloc[:, :5] # ID, SECTION, MODULE, VH, ENSEIGNANT_EXISTANT
    
    # On nettoie la colonne SECTION pour n'afficher que "FILIERE S2" ou "FILIERE S4"
    df.iloc[:, 1] = df.apply(lambda row: f"{f_name} {'S2' if 'S2' in row[1] else 'S4'}", axis=1)
    
    # Remplacer le prof générique "PROF" et les cases vides par "------- (unknown)"
    def clean_teacher(x):
        if pd.isna(x) or str(x).strip() == "":
            return "------- (unknown)"
        # Filtrer le mot clé "PROF" si un vrai prof est aussi sélectionné
        teachers = [t.strip() for t in str(x).split(',') if t.strip() != "PROF" and t.strip()]
        if not teachers:
            return "------- (unknown)"
        return ", ".join(teachers)

    df.iloc[:, 4] = df.iloc[:, 4].apply(clean_teacher)

    filename = os.path.join(OUTPUT_DIR, f"MAQUETTE_A_AFFECTATION_{f_name}.xlsx")
    df.to_excel(filename, index=False, header=False, startrow=4, engine='openpyxl')
    finalize_excel(filename, f"AFFECTATION DES ENSEIGNANTS - {f_name}")
    print(f"✅ Maquette {f_name} générée : {filename}")

def generate_maquette_b(f_name):
    """Génère la fiche répertoire des enseignants pour une filière"""
    print(f"⏳ Génération de la Maquette B (Profs) pour {f_name}...")
    
    columns = ["NOM ET PRÉNOM", "EMAIL", "GSM", "STATUT", "DÉPARTEMENT", "VŒUX / NOTES"]
    df = pd.DataFrame(columns=columns)
    
    # Petit exemple de ligne vide
    df.loc[0] = ["", "", "", "", "", ""]

    filename = os.path.join(OUTPUT_DIR, f"MAQUETTE_B_PROFS_{f_name}.xlsx")
    
    # Export avec style
    df.to_excel(filename, index=False, header=False, startrow=4, engine='openpyxl')
    
    # On réutilisefinalize_excel mais avec 6 colonnes
    wb = load_workbook(filename)
    ws = wb.active
    
    navy_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    # TITRE
    ws.merge_cells('A1:F1')
    ws['A1'] = f"RÉPERTOIRE DES ENSEIGNANTS - {f_name}".upper()
    ws['A1'].font = Font(size=18, bold=True, color="1A365D")
    ws['A1'].alignment = Alignment(horizontal="center")

    # EN-TÊTES
    for i, title in enumerate(columns):
        cell = ws.cell(row=4, column=i+1)
        cell.value = title
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
    
    # LARGEURS
    ws.column_dimensions['A'].width = 35 
    ws.column_dimensions['B'].width = 30 
    ws.column_dimensions['C'].width = 20 
    ws.column_dimensions['D'].width = 15 
    ws.column_dimensions['E'].width = 25 
    ws.column_dimensions['F'].width = 45 

    wb.save(filename)
    print(f"✅ Maquette B {f_name} générée : {filename}")

if __name__ == "__main__":
    # On récupère toutes les filières en base
    with engine.connect() as conn:
        res = conn.execute(text("SELECT name FROM filieres")).fetchall()
        filiere_names = [row[0] for row in res]
    
    for fn in filiere_names:
        generate_maquette_a(fn)
        generate_maquette_b(fn)
