from fastapi import FastAPI, Depends, HTTPException, Body
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import List, Optional
import models, schemas
from sqlalchemy import text
from database import engine, get_db
import sys
import subprocess

models.Base.metadata.create_all(bind=engine)

app = FastAPI(
    title="FST Marrakech Timetabling API",
    version="2.0.0",
    description="API de gestion des emplois du temps — FST de Marrakech (Gueliz)"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000", "http://192.168.56.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ROOT & STATUS

import json, os
from fastapi import Body

# Dossier pour les configurations TP
TP_CONFIG_DIR = "tp_configurations"
if not os.path.exists(TP_CONFIG_DIR):
    os.makedirs(TP_CONFIG_DIR)

@app.post("/save-tp-config/{section_id}")
async def save_tp_config(section_id: int, payload: dict = Body(...)):
    file_path = os.path.join(TP_CONFIG_DIR, f"section_{section_id}.json")
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=4, ensure_ascii=False)
    return {"status": "success", "path": file_path}

@app.get("/load-tp-config/{section_id}")
async def load_tp_config(section_id: int):
    file_path = os.path.join(TP_CONFIG_DIR, f"section_{section_id}.json")
    if not os.path.exists(file_path):
        return []
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return []

@app.get("/")
def read_root():
    return {"message": "FSTM Timetabling API v2.0", "status": "running"}

@app.get("/status")
def get_status():
    return {"status": "ok", "db": "connected"}

@app.get("/stats", response_model=schemas.DashboardStats)
def get_dashboard_stats(db: Session = Depends(get_db)):
    total_teachers = db.query(models.Teacher).count()
    total_rooms = db.query(models.Room).count()
    total_sections = db.query(models.Section).count()
    total_modules = db.query(models.Module).count()
    total_assignments = db.query(models.Assignment).count()
    
    # Calcul des indisponibilités profs
    total_unavail = 0
    teachers = db.query(models.Teacher).all()
    for t in teachers:
        if t.availabilities and "unavailable_slots" in t.availabilities:
            total_unavail += len(t.availabilities["unavailable_slots"])

    return {
        "total_teachers": total_teachers,
        "total_rooms": total_rooms,
        "total_sections": total_sections,
        "total_modules": total_modules,
        "total_assignments": total_assignments,
        "total_teacher_unavailability": total_unavail,
        "hard_violations": 0
    }

@app.get("/audit/master-reference")
def get_master_audit(db: Session = Depends(get_db)):
    """
    Calcule la moyenne globale des 4 indicateurs pédagogiques pour l'emploi du temps Master.
    Retourne un JSON avec les moyennes de Compacité, Pause Midi, Rythme et Pédagogie.
    """
    master = db.query(models.TimetableResult).filter(models.TimetableResult.is_master_reference == True).first()
    if not master:
        return {
            "compacite": 0,
            "pause_dejeuner": 0,
            "rythme_fatigue": 0,
            "pedagogie_cm": 0,
            "status": "No Master Reference"
        }
    
    sections = db.query(models.Section).all()
    if not sections:
        return {"compacite": 0, "pause_dejeuner": 0, "rythme_fatigue": 0, "pedagogie_cm": 0}

    total_metrics = {"compacite": 0.0, "pause_dejeuner": 0.0, "rythme_fatigue": 0.0, "pedagogie_cm": 0.0}
    count = 0

    for sec in sections:
        try:
            # On réutilise la logique interne de audit_section
            audit = audit_section(sec.id, mode=master.algo_type, db=db)
            if audit and "details" in audit:
                total_metrics["compacite"] += audit["details"].get("compacite", 0)
                total_metrics["pause_dejeuner"] += audit["details"].get("pause_dejeuner", 0)
                total_metrics["rythme_fatigue"] += audit["details"].get("rythme_fatigue", 0)
                total_metrics["pedagogie_cm"] += audit["details"].get("pedagogie_cm", 0)
                count += 1
        except:
            continue

    if count == 0:
        return {"compacite": 0, "pause_dejeuner": 0, "rythme_fatigue": 0, "pedagogie_cm": 0}

    return {
        "compacite": round(total_metrics["compacite"] / count, 1),
        "pause_dejeuner": round(total_metrics["pause_dejeuner"] / count, 1),
        "rythme_fatigue": round(total_metrics["rythme_fatigue"] / count, 1),
        "pedagogie_cm": round(total_metrics["pedagogie_cm"] / count, 1)
    }


# TEACHERS  /teachers

@app.get("/teachers", response_model=List[schemas.Teacher])
def list_teachers(skip: int = 0, limit: int = 500, db: Session = Depends(get_db)):
    return db.query(models.Teacher).offset(skip).limit(limit).all()

@app.get("/teachers/{teacher_id}", response_model=schemas.Teacher)
def get_teacher(teacher_id: int, db: Session = Depends(get_db)):
    t = db.query(models.Teacher).filter(models.Teacher.id == teacher_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Enseignant introuvable")
    return t

@app.post("/teachers", response_model=schemas.Teacher, status_code=201)
def create_teacher(teacher: schemas.TeacherCreate, db: Session = Depends(get_db)):
    existing = db.query(models.Teacher).filter(models.Teacher.email == teacher.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email déjà utilisé")
    db_t = models.Teacher(**teacher.model_dump())
    db.add(db_t); db.commit(); db.refresh(db_t)
    return db_t

@app.put("/teachers/{teacher_id}", response_model=schemas.Teacher)
def update_teacher(teacher_id: int, data: schemas.TeacherUpdate, db: Session = Depends(get_db)):
    t = db.query(models.Teacher).filter(models.Teacher.id == teacher_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Enseignant introuvable")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(t, k, v)
    db.commit(); db.refresh(t)
    return t

@app.delete("/teachers/{teacher_id}", status_code=204)
def delete_teacher(teacher_id: int, db: Session = Depends(get_db)):
    t = db.query(models.Teacher).filter(models.Teacher.id == teacher_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Enseignant introuvable")
    db.delete(t); db.commit()


# ROOMS  /rooms

@app.get("/rooms", response_model=List[schemas.Room])
def list_rooms(type: str = None, db: Session = Depends(get_db)):
    q = db.query(models.Room)
    if type:
        q = q.filter(models.Room.type == type)
    return q.all()

@app.get("/rooms/{room_id}", response_model=schemas.Room)
def get_room(room_id: int, db: Session = Depends(get_db)):
    r = db.query(models.Room).filter(models.Room.id == room_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Salle introuvable")
    return r

@app.post("/rooms", response_model=schemas.Room, status_code=201)
def create_room(room: schemas.RoomCreate, db: Session = Depends(get_db)):
    db_r = models.Room(**room.model_dump())
    db.add(db_r); db.commit(); db.refresh(db_r)
    return db_r

@app.put("/rooms/{room_id}", response_model=schemas.Room)
def update_room(room_id: int, data: schemas.RoomUpdate, db: Session = Depends(get_db)):
    r = db.query(models.Room).filter(models.Room.id == room_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Salle introuvable")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(r, k, v)
    db.commit(); db.refresh(r)
    return r

@app.delete("/rooms/{room_id}", status_code=204)
def delete_room(room_id: int, db: Session = Depends(get_db)):
    r = db.query(models.Room).filter(models.Room.id == room_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Salle introuvable")
    db.delete(r); db.commit()


# GROUPE FILIERES  /groupe-filieres

@app.get("/groupe-filieres", response_model=List[schemas.GroupeFiliere])
def list_groupe_filieres(db: Session = Depends(get_db)):
    return db.query(models.GroupeFiliere).all()

@app.post("/groupe-filieres", response_model=schemas.GroupeFiliere, status_code=201)
def create_groupe_filiere(group: schemas.GroupeFiliereCreate, db: Session = Depends(get_db)):
    db_g = models.GroupeFiliere(**group.model_dump())
    db.add(db_g); db.commit(); db.refresh(db_g)
    return db_g

@app.delete("/groupe-filieres/{group_id}", status_code=204)
def delete_groupe_filiere(group_id: int, db: Session = Depends(get_db)):
    g = db.query(models.GroupeFiliere).filter(models.GroupeFiliere.id == group_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Groupe introuvable")
    db.delete(g); db.commit()

@app.put("/groupe-filieres/{group_id}", response_model=schemas.GroupeFiliere)
def update_groupe_filiere(group_id: int, data: schemas.GroupeFiliereCreate, db: Session = Depends(get_db)):
    g = db.query(models.GroupeFiliere).filter(models.GroupeFiliere.id == group_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Groupe introuvable")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(g, k, v)
    db.commit(); db.refresh(g)
    return g


# SECTIONS  /sections

@app.get("/sections", response_model=List[schemas.Section])
def list_sections(db: Session = Depends(get_db)):
    return db.query(models.Section).all()

@app.post("/sections", response_model=schemas.Section, status_code=201)
def create_section(data: schemas.SectionCreate, db: Session = Depends(get_db)):
    new_s = models.Section(
        name=data.name,
        semestre=data.semestre,
        total_capacity=data.total_capacity
    )
    for g_id in data.groupe_ids:
        g = db.query(models.GroupeFiliere).filter(models.GroupeFiliere.id == g_id).first()
        if g:
            new_s.groupes.append(g)
    db.add(new_s); db.commit(); db.refresh(new_s)
    return new_s

@app.delete("/sections/{section_id}", status_code=204)
def delete_section(section_id: int, db: Session = Depends(get_db)):
    s = db.query(models.Section).filter(models.Section.id == section_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Section introuvable")
    db.delete(s); db.commit()

@app.put("/sections/{section_id}", response_model=schemas.Section)
def update_section(section_id: int, data: schemas.SectionCreate, db: Session = Depends(get_db)):
    s = db.query(models.Section).filter(models.Section.id == section_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Section introuvable")
    
    s.name = data.name
    s.semestre = data.semestre
    s.total_capacity = data.total_capacity
    
    # Mise à jour des groupes associés (pour les sections c'est souvent les cohortes parentes)
    s.groupes = []
    for g_id in data.groupe_ids:
        g = db.query(models.GroupeFiliere).filter(models.GroupeFiliere.id == g_id).first()
        if g:
            s.groupes.append(g)
            
    db.commit(); db.refresh(s)
    return s


@app.get("/sections/{section_id}/tp-blocking")
def get_section_tp_blocking(section_id: int, db: Session = Depends(get_db)):
    """
    Renvoie les IDs des créneaux bloqués pour chaque groupe de la section.
    Priorité aux règles en DB, sinon fallback sur la logique par défaut (Section 1).
    """
    section = db.query(models.Section).filter(models.Section.id == section_id).first()
    if not section:
        raise HTTPException(status_code=404, detail="Section non trouvée")
    
    timeslots = db.query(models.Timeslot).all()
    groups = db.query(models.TDGroup).filter(models.TDGroup.section_id == section_id).all()
    group_ids = [g.id for g in groups]

    # Tentative de récupération des règles en DB
    rules = db.query(models.TPSanctuarization).filter(models.TPSanctuarization.group_id.in_(group_ids)).all()
    
    result = {}
    
    if not rules:
        # LOGIQUE PAR DÉFAUT (Section 1 uniquement)
        for g in groups:
            import re
            m = re.search(r'Gr\s*(\d+)', g.name)
            if not m:
                result[g.id] = []
                continue
            g_num = int(m.group(1))
            blocked_ids = []
            for s in timeslots:
                h = s.start_time.hour
                is_morning = (h < 13)
                conflict = False
                if section_id == 1:
                    if g_num in [1, 2] and (s.day.upper() in ["MARDI", "JEUDI"] and not is_morning): conflict = True
                    elif g_num in [3, 4] and (s.day.upper() in ["MERCREDI", "VENDREDI"] and is_morning): conflict = True
                    elif g_num in [5, 6] and (s.day.upper() in ["LUNDI", "MERCREDI"] and is_morning): conflict = True
                if conflict: blocked_ids.append(s.id)
            result[g.id] = blocked_ids
    else:
        # LOGIQUE DYNAMIQUE (Depuis DB)
        for g in groups:
            g_rules = [r for r in rules if r.group_id == g.id]
            blocked_ids = []
            for r in g_rules:
                for s in timeslots:
                    h = s.start_time.hour
                    is_morning = (h < 13)
                    if s.day.upper() == r.day.upper() and is_morning == r.is_morning:
                        blocked_ids.append(s.id)
            result[g.id] = blocked_ids

    return result

@app.get("/sections/{section_id}/sanctuarizations", response_model=List[schemas.TPSanctuarization])
def list_section_sanctuarizations(section_id: int, db: Session = Depends(get_db)):
    """Renvoie les règles brutes pour l'interface de config."""
    groups = db.query(models.TDGroup).filter(models.TDGroup.section_id == section_id).all()
    group_ids = [g.id for g in groups]
    return db.query(models.TPSanctuarization).filter(models.TPSanctuarization.group_id.in_(group_ids)).all()

@app.post("/sections/{section_id}/sanctuarizations")
def update_section_sanctuarizations(section_id: int, data: schemas.TPBulkUpdate, db: Session = Depends(get_db)):
    """Mise à jour en masse des règles de sanctuarisation."""
    groups = db.query(models.TDGroup).filter(models.TDGroup.section_id == section_id).all()
    group_ids = [g.id for g in groups]
    
    # 1. Nettoyage
    db.query(models.TPSanctuarization).filter(models.TPSanctuarization.group_id.in_(group_ids)).delete(synchronize_session=False)
    
    # 2. Insertion
    for r in data.rules:
        new_rule = models.TPSanctuarization(
            group_id=r.group_id,
            day=r.day.upper(),
            is_morning=r.is_morning
        )
        db.add(new_rule)
    
    db.commit()
    return {"message": "Configuration mise à jour avec succès"}

@app.get("/sections/{section_id}/tp-blocking")
def get_section_tp_blocking(section_id: int, db: Session = Depends(get_db)):
    """
    Retourne un dict {group_id: [timeslot_ids]} représentant
    les créneaux bloqués par la sanctuarisation TP pour chaque groupe.
    Utilisé par le frontend pour colorier la grille de gestion des TP.
    """
    # 1. Récupérer les groupes de la section
    groups = db.query(models.TDGroup).filter(models.TDGroup.section_id == section_id).all()
    group_ids = [g.id for g in groups]

    # 2. Récupérer les règles de sanctuarisation
    rules = db.query(models.TPSanctuarization).filter(
        models.TPSanctuarization.group_id.in_(group_ids)
    ).all()

    # 3. Récupérer tous les timeslots
    timeslots = db.query(models.TimeSlot).all()

    # 4. Pour chaque règle, trouver les timeslots correspondants
    # Matin = 08:30 et 10:35 | Après-midi = 14:30 et 16:35
    morning_starts = ["08:30", "10:35"]
    afternoon_starts = ["14:30", "16:35"]

    result: dict = {gid: [] for gid in group_ids}

    for rule in rules:
        target_starts = morning_starts if rule.is_morning else afternoon_starts
        rule_day = rule.day.upper()

        for ts in timeslots:
            ts_day = ts.day.upper()
            ts_start = ts.start_time[:5]  # "HH:MM"

            if ts_day == rule_day and ts_start in target_starts:
                if rule.group_id in result:
                    result[rule.group_id].append(ts.id)

    return result

@app.get("/tp-sanctuarizations-numbers")
def get_tp_blocking_numbers(db: Session = Depends(get_db)):
    count = db.query(models.TPSanctuarization).count()
    return {"count_tp": count}



# TD GROUPS  /td-groups

@app.get("/td-groups", response_model=List[schemas.TDGroup])
def list_td_groups(db: Session = Depends(get_db)):
    return db.query(models.TDGroup).all()

@app.post("/td-groups", response_model=schemas.TDGroup, status_code=201)
def create_td_group(group: schemas.TDGroupCreate, db: Session = Depends(get_db)):
    db_g = models.TDGroup(**group.model_dump())
    db.add(db_g); db.commit(); db.refresh(db_g)
    return db_g

@app.delete("/td-groups/{tdgroup_id}", status_code=204)
def delete_td_group(tdgroup_id: int, db: Session = Depends(get_db)):
    g = db.query(models.TDGroup).filter(models.TDGroup.id == tdgroup_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="TDGroup introuvable")
    db.delete(g); db.commit()

@app.put("/td-groups/{tdgroup_id}", response_model=schemas.TDGroup)
def update_td_group(tdgroup_id: int, data: schemas.TDGroupCreate, db: Session = Depends(get_db)):
    g = db.query(models.TDGroup).filter(models.TDGroup.id == tdgroup_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="TDGroup introuvable")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(g, k, v)
    db.commit(); db.refresh(g)
    return g


# MODULES  /modules

@app.get("/modules", response_model=List[schemas.Module])
def list_modules(db: Session = Depends(get_db)):
    return db.query(models.Module).all()

@app.post("/modules", response_model=schemas.Module, status_code=201)
def create_module(module: schemas.ModuleCreate, db: Session = Depends(get_db)):
    existing = db.query(models.Module).filter(models.Module.code == module.code).first()
    if existing:
        raise HTTPException(status_code=400, detail="Code module déjà utilisé")
    db_m = models.Module(**module.model_dump())
    db.add(db_m); db.commit(); db.refresh(db_m)
    return db_m

@app.put("/modules/{module_id}", response_model=schemas.Module)
def update_module(module_id: int, data: schemas.ModuleUpdate, db: Session = Depends(get_db)):
    m = db.query(models.Module).filter(models.Module.id == module_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Module introuvable")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(m, k, v)
    db.commit(); db.refresh(m)
    return m

@app.delete("/modules/{module_id}", status_code=204)
def delete_module(module_id: int, db: Session = Depends(get_db)):
    m = db.query(models.Module).filter(models.Module.id == module_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Module introuvable")
    db.delete(m); db.commit()


# GROUPE MODULES  /groupe-modules

@app.get("/groupe-modules", response_model=List[schemas.GroupeModule])
def list_groupe_modules(db: Session = Depends(get_db)):
    return db.query(models.GroupeModule).all()

@app.post("/groupe-modules", response_model=schemas.GroupeModule, status_code=201)
def create_groupe_module(data: schemas.GroupeModuleCreate, db: Session = Depends(get_db)):
    existing = db.query(models.GroupeModule).filter(models.GroupeModule.module_id == data.module_id).first()
    if existing: raise HTTPException(400, "Ce module a déjà un groupe d'effectif assigné.")
    db_gm = models.GroupeModule(module_id=data.module_id, effectif=data.effectif)
    for gid in data.groupe_ids:
        g = db.query(models.GroupeFiliere).filter(models.GroupeFiliere.id == gid).first()
        if g: db_gm.groupes.append(g)
    db.add(db_gm); db.commit(); db.refresh(db_gm)
    return db_gm

@app.put("/groupe-modules/{gm_id}", response_model=schemas.GroupeModule)
def update_groupe_module(gm_id: int, data: schemas.GroupeModuleUpdate, db: Session = Depends(get_db)):
    db_gm = db.query(models.GroupeModule).filter(models.GroupeModule.id == gm_id).first()
    if not db_gm: raise HTTPException(404, "GroupeModule introuvable")
    if data.effectif is not None: db_gm.effectif = data.effectif
    if data.groupe_ids is not None:
        db_gm.groupes = []
        for gid in data.groupe_ids:
            g = db.query(models.GroupeFiliere).filter(models.GroupeFiliere.id == gid).first()
            if g: db_gm.groupes.append(g)
    db.commit(); db.refresh(db_gm)
    return db_gm

@app.delete("/groupe-modules/{gm_id}", status_code=204)
def delete_groupe_module(gm_id: int, db: Session = Depends(get_db)):
    db_gm = db.query(models.GroupeModule).filter(models.GroupeModule.id == gm_id).first()
    if not db_gm: raise HTTPException(404, "GroupeModule introuvable")
    db.delete(db_gm); db.commit()


# MODULE PARTS  /module-parts

@app.get("/module-parts", response_model=List[schemas.ModulePart])
def list_module_parts(db: Session = Depends(get_db)):
    return db.query(models.ModulePart).all()

@app.post("/module-parts", response_model=schemas.ModulePart, status_code=201)
def create_module_part(part: schemas.ModulePartCreate, db: Session = Depends(get_db)):
    db_p = models.ModulePart(**part.model_dump())
    db.add(db_p); db.commit(); db.refresh(db_p)
    return db_p

@app.put("/module-parts/{part_id}", response_model=schemas.ModulePart)
def update_module_part(part_id: int, part: schemas.ModulePartUpdate, db: Session = Depends(get_db)):
    db_p = db.query(models.ModulePart).filter(models.ModulePart.id == part_id).first()
    if not db_p: raise HTTPException(404, "Partie de module introuvable")
    data = part.model_dump(exclude_unset=True)
    for k,v in data.items(): setattr(db_p, k, v)
    db.commit(); db.refresh(db_p)
    return db_p

@app.delete("/module-parts/{part_id}", status_code=204)
def delete_module_part(part_id: int, db: Session = Depends(get_db)):
    db_p = db.query(models.ModulePart).filter(models.ModulePart.id == part_id).first()
    if not db_p: raise HTTPException(404, "Partie de module introuvable")
    db.delete(db_p); db.commit()


# FILIERES  /filieres

@app.get("/filieres", response_model=List[schemas.Filiere])
def list_filieres(db: Session = Depends(get_db)):
    return db.query(models.Filiere).all()

@app.get("/filieres/{filiere_id}/stats")
def get_filiere_stats(filiere_id: int, db: Session = Depends(get_db)):
    """Retourne les stats dynamiques d'une filière pour les KPI cards"""
    try:
        f = db.query(models.Filiere).filter(models.Filiere.id == filiere_id).first()
        if not f:
            raise HTTPException(404, "Filière introuvable")

        # Compter les sections de cette filière
        sections_query = text("""
            SELECT DISTINCT s.id, s.name
            FROM sections s
            JOIN section_groupes sg ON sg.section_id = s.id
            JOIN groupe_filieres gf ON gf.id = sg.groupe_id
            WHERE gf.filiere_id = :fid
        """)
        sections = db.execute(sections_query, {"fid": filiere_id}).fetchall()
        section_ids = [s.id for s in sections]

        # Compter les modules distincts enseignés dans cette filière
        module_count = 0
        if section_ids:
            module_query = text("""
                SELECT COUNT(DISTINCT m.id) as nb
                FROM modules m
                JOIN module_parts mp ON mp.module_id = m.id
                JOIN assignments a ON a.module_part_id = mp.id
                WHERE a.section_id = ANY(:sids)
            """)
            result = db.execute(module_query, {"sids": section_ids}).fetchone()
            module_count = result.nb if result else 0

        # Compter les groupes TD (effectif = nb_groupes × 30 étudiants environ)
        td_count = 0
        if section_ids:
            td_query = text("""
                SELECT COUNT(id) as nb FROM td_groups
                WHERE section_id = ANY(:sids)
            """)
            result2 = db.execute(td_query, {"sids": section_ids}).fetchone()
            td_count = result2.nb if result2 else 0

        effectif_estime = td_count * 30  # ~30 étudiants par groupe TD

        return {
            "filiere_id": filiere_id,
            "filiere_name": f.name,
            "nb_sections": len(sections),
            "sections": [{"id": s.id, "name": s.name} for s in sections],
            "nb_modules": module_count,
            "nb_td_groups": td_count,
            "effectif_estime": effectif_estime if effectif_estime > 0 else 0,
        }
    except Exception as e:
        print(f"Error in get_filiere_stats: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Backend Error: {str(e)}")

@app.post("/filieres", response_model=schemas.Filiere, status_code=201)
def create_filiere(f: schemas.FiliereCreate, db: Session = Depends(get_db)):
    db_f = models.Filiere(**f.model_dump())
    db.add(db_f); db.commit(); db.refresh(db_f)
    return db_f

@app.put("/filieres/{filiere_id}", response_model=schemas.Filiere)
def update_filiere(filiere_id: int, data: schemas.FiliereUpdate, db: Session = Depends(get_db)):
    f = db.query(models.Filiere).filter(models.Filiere.id == filiere_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Filière introuvable")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(f, k, v)
    db.commit(); db.refresh(f)
    return f

@app.delete("/filieres/{filiere_id}", status_code=204)
def delete_filiere(filiere_id: int, db: Session = Depends(get_db)):
    f = db.query(models.Filiere).filter(models.Filiere.id == filiere_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Filière introuvable")
    db.delete(f); db.commit()


# DEPARTMENTS  /departments

@app.get("/departments", response_model=List[schemas.Department])
def list_departments(db: Session = Depends(get_db)):
    return db.query(models.Department).all()


# TIMESLOTS  /timeslots

@app.get("/timeslots", response_model=List[schemas.Timeslot])
def list_timeslots(db: Session = Depends(get_db)):
    return db.query(models.Timeslot).all()

@app.post("/timeslots", response_model=schemas.Timeslot, status_code=201)
def create_timeslot(slot: schemas.TimeslotCreate, db: Session = Depends(get_db)):
    db_s = models.Timeslot(**slot.model_dump())
    db.add(db_s); db.commit(); db.refresh(db_s)
    return db_s

@app.put("/timeslots/{slot_id}", response_model=schemas.Timeslot)
def update_timeslot(slot_id: int, data: schemas.TimeslotUpdate, db: Session = Depends(get_db)):
    s = db.query(models.Timeslot).filter(models.Timeslot.id == slot_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Créneau introuvable")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(s, k, v)
    db.commit(); db.refresh(s)
    return s

@app.delete("/timeslots/{slot_id}", status_code=204)
def delete_timeslot(slot_id: int, db: Session = Depends(get_db)):
    s = db.query(models.Timeslot).filter(models.Timeslot.id == slot_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Créneau introuvable")
    db.delete(s); db.commit()


# ASSIGNMENTS  /assignments

@app.get("/assignments", response_model=List[schemas.Assignment])
def list_assignments(db: Session = Depends(get_db)):
    return db.query(models.Assignment).all()

@app.post("/assignments", response_model=schemas.Assignment, status_code=201)
def create_assignment(data: schemas.AssignmentCreate, db: Session = Depends(get_db)):
    new_a = models.Assignment(
        module_part_id=data.module_part_id,
        teacher_id=data.teacher_id,
        room_id=data.room_id,
        slot_id=data.slot_id,
        section_id=data.section_id,
        is_locked=data.is_locked,
        tp_subgroup=data.tp_subgroup,
        alternance=data.alternance
    )
    # Ajout des groupes TD si c'est un TD/TP
    for g_id in data.tdgroup_ids:
        g = db.query(models.TDGroup).filter(models.TDGroup.id == g_id).first()
        if g:
            new_a.td_groups.append(g)
    
    db.add(new_a); db.commit(); db.refresh(new_a)
    return new_a

@app.delete("/assignments/{assignment_id}", status_code=204)
def delete_assignment(assignment_id: int, db: Session = Depends(get_db)):
    a = db.query(models.Assignment).filter(models.Assignment.id == assignment_id).first()
    if not a:
        raise HTTPException(status_code=404, detail="Affectation introuvable")
    db.delete(a); db.commit()

@app.put("/assignments/{assignment_id}", response_model=schemas.Assignment)
def update_assignment(assignment_id: int, data: schemas.AssignmentUpdate, db: Session = Depends(get_db)):
    db_a = db.query(models.Assignment).filter(models.Assignment.id == assignment_id).first()
    if not db_a:
        raise HTTPException(status_code=404, detail="Affectation introuvable")
    
    # Mise à jour sélective
    update_data = data.model_dump(exclude_unset=True)
    
    if "module_part_id" in update_data: db_a.module_part_id = update_data["module_part_id"]
    if "teacher_id" in update_data: db_a.teacher_id = update_data["teacher_id"]
    if "section_id" in update_data: db_a.section_id = update_data["section_id"]
    if "is_locked" in update_data: db_a.is_locked = update_data["is_locked"]
    if "room_id" in update_data: db_a.room_id = update_data["room_id"]
    if "slot_id" in update_data: db_a.slot_id = update_data["slot_id"]
    if "tp_subgroup" in update_data: db_a.tp_subgroup = update_data["tp_subgroup"]
    if "alternance" in update_data: db_a.alternance = update_data["alternance"]

    # Mise à jour des groupes TD si spécifiés
    if "tdgroup_ids" in update_data:
        db_a.td_groups = []
        for g_id in update_data["tdgroup_ids"]:
            g = db.query(models.TDGroup).filter(models.TDGroup.id == g_id).first()
            if g:
                db_a.td_groups.append(g)

    db.commit()
    db.refresh(db_a)
    return db_a



# --- DISPONIBILITÉS EN TEMPS RÉEL (EDITION MANUELLE) ---

@app.get("/available-resources")
def get_available_resources(slot_id: int, assignment_id: Optional[int] = None, db: Session = Depends(get_db)):
    """
    Renvoie les listes filtrées de professeurs et de salles disponibles pour un créneau donné.
    Si assignment_id est fourni, on ignore l'occupation par cette affectation elle-même.
    """
    # 1. Trouver les occupations actuelles sur ce créneau
    occupied_query = db.query(models.Assignment).filter(models.Assignment.slot_id == slot_id)
    if assignment_id:
        occupied_query = occupied_query.filter(models.Assignment.id != assignment_id)
    
    occupations = occupied_query.all()
    occupied_room_ids = {a.room_id for a in occupations if a.room_id}
    occupied_teacher_ids = {a.teacher_id for a in occupations if a.teacher_id}

    # 2. Chambres Libres
    all_rooms = db.query(models.Room).all()
    available_rooms = [r for r in all_rooms if r.id not in occupied_room_ids]

    # 3. Professeurs Libres et Disponibles
    all_teachers = db.query(models.Teacher).all()
    slot = db.query(models.Timeslot).filter(models.Timeslot.id == slot_id).first()
    
    available_teachers = []
    if slot:
        day_name = slot.day.capitalize() # ex: "Lundi"
        slot_str = f"{slot.start_time.strftime('%H:%M')}-{slot.end_time.strftime('%H:%M')}"
        
        for t in all_teachers:
            # Ne doit pas être occupé ailleurs
            if t.id in occupied_teacher_ids:
                continue
            
            # Vérification des préférences (Optionnel : si renseigné)
            if t.availabilities:
                # Format attendu : {"Lundi": ["08:30-10:25", ...]}
                day_slots = t.availabilities.get(day_name, [])
                if day_slots and slot_str not in day_slots:
                    # Si le prof a renseigné des créneaux mais que celui-ci n'y est pas
                    # Note: on peut être plus souple si availabilities est vide
                    continue
            
            available_teachers.append(t)
    else:
        # Fallback si slot non trouvé
        available_teachers = [t for t in all_teachers if t.id not in occupied_teacher_ids]

    return {
        "available_rooms": available_rooms,
        "available_teachers": available_teachers
    }


# --- API POUR L'ALGORITHME LOCAL ---

@app.get("/data-for-solver")
def get_data_for_solver(db: Session = Depends(get_db)):
    """
    Exporte toutes les données de la base SQL en JSON 
    pour que l'algorithme local sur Windows puisse les traiter.
    """
    return {
        "rooms": db.query(models.Room).all(),
        "sections": db.query(models.Section).all(),
        "td_groups": db.query(models.TDGroup).all(),
        "timeslots": db.query(models.Timeslot).all(),
        "modules": db.query(models.Module).all(),
        # On peut ajouter ici d'autres ressources si besoin
    }

@app.post("/save-assignments", status_code=201)
def save_assignments(assignments: List[schemas.AssignmentCreate], db: Session = Depends(get_db)):
    """
    Reçoit le résultat de l'IA (le planning) et l'enregistre dans la base.
    """
    try:
        # 1. On vide proprement l'ancien planning
        # On utilise une boucle pour être sûr que les relations M:M sont nettoyées
        old_assignments = db.query(models.Assignment).all()
        for old in old_assignments:
            old.td_groups = [] # Nettoyer la table d'association
            db.delete(old)
        db.flush() 

        # 2. On enregistre le nouveau planning
        for a in assignments:
            new_a = models.Assignment(
                module_part_id=a.module_part_id,
                teacher_id=a.teacher_id,
                room_id=a.room_id,
                slot_id=a.slot_id,
                section_id=a.section_id,
                is_locked=a.is_locked,
                tp_subgroup=a.tp_subgroup,
                alternance=a.alternance
            )
            # Ajout des groupes TD si présents
            if a.tdgroup_ids:
                for g_id in a.tdgroup_ids:
                    g = db.query(models.TDGroup).filter(models.TDGroup.id == g_id).first()
                    if g:
                        new_a.td_groups.append(g)
            
            db.add(new_a)
        
        db.commit()
        return {"message": "Planning sauvegardé en base de données avec tous les détails."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur lors de la sauvegarde : {str(e)}")

import os
import json

@app.get("/preview-schedule")
def get_preview_schedule(mode: str = "alns", db: Session = Depends(get_db)):
    """Récupère le dernier résultat BRUT (is_validated=False) pour un mode donné."""
    result = db.query(models.TimetableResult)\
               .filter(models.TimetableResult.algo_type == mode, models.TimetableResult.is_validated == False)\
               .order_by(models.TimetableResult.id.desc())\
               .first()
    
    if not result:
        return []
        
    return result.data

from typing import Optional

@app.post("/commit-preview", status_code=201)
def commit_preview(mode: str = "alns", result_id: Optional[int] = None, db: Session = Depends(get_db)):
    """
    1. Marque le dernier résultat (ou celui par id) comme 'validé' (Archive).
    2. Copie les données du JSON vers la table 'assignments' pour édition manuelle.
    """
    try:
        if result_id:
            res = db.query(models.TimetableResult).filter(models.TimetableResult.id == result_id).first()
        else:
            res = db.query(models.TimetableResult)\
                    .filter(models.TimetableResult.algo_type == mode)\
                    .order_by(models.TimetableResult.id.desc())\
                    .first()
        
        if not res:
            raise HTTPException(status_code=404, detail="Aucun résultat trouvé")

        # 3. RÉINITIALISATION ET MISE À JOUR DE LA TABLE DE TRAVAIL (assignments)
        # Étape A : On "vide" le planning actuel et les groupes associés
        all_assignments = db.query(models.Assignment).all()
        for a in all_assignments:
            a.slot_id = None
            a.room_id = None
            a.td_groups = [] # On vide les groupes pour les ré-associer proprement
        db.flush()

        # Étape B : On applique les nouvelles positions et on relie les groupes
        for item in res.data:
            assignment_id = item.get('id')
            if not assignment_id: continue

            db_a = db.query(models.Assignment).filter(models.Assignment.id == assignment_id).first()
            
            if not db_a:
                # Si par hasard l'affectation n'existe pas en base, on la crée
                db_a = models.Assignment(id=assignment_id, module_part_id=item['module_part_id'])
                db.add(db_a)

            # Mise à jour des données de base
            db_a.room_id = item.get('room_id')
            db_a.slot_id = item.get('slot_id')
            db_a.teacher_id = item.get('teacher_id')
            db_a.is_locked = item.get('is_locked', False)
            db_a.section_id = item.get('section_id')
            db_a.tp_subgroup = item.get('tp_subgroup')
            db_a.alternance = item.get('alternance')

            # --- RESTAURATION DES GROUPES TD (CRUCIAL POUR LE SCORE) ---
            group_data = item.get('td_groups', [])
            for g_item in group_data:
                # Le JSON peut contenir soit l'ID directement, soit l'objet groupe
                g_id = g_item['id'] if isinstance(g_item, dict) else g_item
                g_obj = db.query(models.TDGroup).filter(models.TDGroup.id == g_id).first()
                if g_obj:
                    db_a.td_groups.append(g_obj)

        db.commit()


        return {"status": "success", "message": f"Résultat {mode} activé et copié dans l'éditeur interactif."}
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur lors du commit : {str(e)}")

@app.post("/save-manual-session", status_code=201)
def save_manual_session(
    name: Optional[str] = "Version Manuelle", 
    edit_id: Optional[int] = None, 
    score_hard: Optional[int] = 0,
    score_soft: Optional[float] = 0.0,
    algo_type: Optional[str] = "manual",
    db: Session = Depends(get_db)
):
    """
    Prend le planning actuel dans la table 'assignments', 
    génère un snapshot JSON et le sauvegarde dans 'timetable_results'.
    Si edit_id est fourni, met à jour la version existante, sinon crée une nouvelle version validée.
    """
    from datetime import datetime
    try:
        # 1. On récupère TOUTES les affectations actuelles
        all_asgn = db.query(models.Assignment).all()
        
        # 2. On prépare le JSON de données
        json_data = []
        hard_conflicts = 0
        
        # On va aussi charger les timeslots pour les calculs de conflits
        for a in all_asgn:
            # Conversion pour le JSON
            json_data.append({
                "id": a.id,
                "module_part_id": a.module_part_id,
                "teacher_id": a.teacher_id,
                "room_id": a.room_id,
                "slot_id": a.slot_id,
                "section_id": a.section_id,
                "is_locked": a.is_locked,
                "tp_subgroup": a.tp_subgroup,
                "alternance": a.alternance,
                "td_groups": [{"id": g.id} for g in a.td_groups]
            })
            
            # 3. Vérification basique des conflits Hard (Doubles réservations)
            if a.slot_id:
                # Professeur (sauf prof 'A Déterminer' ID 231)
                prof_conflict = db.query(models.Assignment).filter(
                    models.Assignment.id != a.id,
                    models.Assignment.slot_id == a.slot_id,
                    models.Assignment.teacher_id == a.teacher_id,
                    models.Assignment.teacher_id != 231
                ).first()
                if prof_conflict: hard_conflicts += 1
                
                # Salle
                if a.room_id:
                    room_conflict = db.query(models.Assignment).filter(
                        models.Assignment.id != a.id,
                        models.Assignment.slot_id == a.slot_id,
                        models.Assignment.room_id == a.room_id
                    ).first()
                    if room_conflict: hard_conflicts += 1

        if hard_conflicts > 0:
             raise HTTPException(status_code=400, detail=f"Le planning contient {hard_conflicts} conflits majeurs. Merci de les corriger avant de sauvegarder.")

        # 4. Enregistrement en BD (UPDATE si edit_id, sinon INSERT)
        validated_res = None
        if edit_id:
            validated_res = db.query(models.TimetableResult).filter(models.TimetableResult.id == edit_id).first()
            
        if not validated_res:
             validated_res = models.TimetableResult(algo_type=algo_type, is_validated=True)
             db.add(validated_res)

        if name:
             validated_res.name = name
        elif not validated_res.name:
             validated_res.name = "Version Manuelle"
             
        validated_res.created_at = datetime.now().isoformat()
        validated_res.data = json_data
        validated_res.score_hard = score_hard
        validated_res.score_soft = score_soft
        validated_res.is_validated = True
        
        # LOGIQUE AUTO-MASTER : Si aucun master n'existe, on le devient
        existing_master = db.query(models.TimetableResult).filter(models.TimetableResult.is_master_reference == True).first()
        if not existing_master:
            validated_res.is_master_reference = True
        
        db.commit()
        
        return {"status": "success", "message": "Planning de production mis à jour avec succès."}
        
    except HTTPException as he:
        raise he
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur lors de la sauvegarde : {str(e)}")


# --- GESTION DES RÉSULTATS D'OPTIMISATION ---

@app.get("/timetable-results/{res_id}", response_model=schemas.TimetableResult)
def get_timetable_result(res_id: int, db: Session = Depends(get_db)):
    """Récupère un résultat spécifique par son ID."""
    res = db.query(models.TimetableResult).filter(models.TimetableResult.id == res_id).first()
    if not res:
        raise HTTPException(status_code=404, detail="Résultat introuvable")
    return res

@app.get("/timetable-results", response_model=List[schemas.TimetableResult])
def list_timetable_results(db: Session = Depends(get_db)):
    """Liste tous les emplois du temps générés par ordre chronologique inverse."""
    return db.query(models.TimetableResult).order_by(models.TimetableResult.id.desc()).all()

@app.post("/timetable-results", response_model=schemas.TimetableResult, status_code=201)
def create_timetable_result(result: schemas.TimetableResultCreate, db: Session = Depends(get_db)):
    """Sauvegarde un nouveau résultat de génération (appelé par le solveur)."""
    # Si aucun master, le nouveau le devient
    existing_master = db.query(models.TimetableResult).filter(models.TimetableResult.is_master_reference == True).first()
    
    db_res = models.TimetableResult(**result.model_dump())
    if not existing_master:
        db_res.is_master_reference = True
        
    db.add(db_res); db.commit(); db.refresh(db_res)
    return db_res

@app.get("/timetable-results/master/reference")
def get_master_reference(db: Session = Depends(get_db)):
    """Récupère l'EDT marqué comme Master, ou le dernier résultat 'fused' par défaut."""
    # 1. Tentative : Master explicite
    res = db.query(models.TimetableResult).filter(models.TimetableResult.is_master_reference == True).first()
    
    # 2. Fallback : Dernier résultat de type 'fused' (IA)
    if not res:
        res = db.query(models.TimetableResult).filter(
            models.TimetableResult.algo_type == "fused"
        ).order_by(models.TimetableResult.id.desc()).first()
        
    if not res:
        raise HTTPException(status_code=404, detail="Aucun emploi du temps (Master ou Fused) trouvé en base de données.")
        
    return res

@app.post("/timetable-results/{res_id}/set-master")
def set_master_reference(res_id: int, db: Session = Depends(get_db)):
    """Définit un emploi du temps comme référence master et désactive les autres."""
    # 1. On désactive l'ancien master
    db.query(models.TimetableResult).update({models.TimetableResult.is_master_reference: False})
    
    # 2. On active le nouveau
    res = db.query(models.TimetableResult).filter(models.TimetableResult.id == res_id).first()
    if not res:
        db.rollback()
        raise HTTPException(status_code=404, detail="Résultat introuvable")
    
    res.is_master_reference = True
    db.commit()
    return {"message": f"'{res.name}' est désormais l'emploi du temps de référence officiel."}

@app.put("/timetable-results/{res_id}/validate", response_model=schemas.TimetableResult)
def validate_timetable_result(res_id: int, db: Session = Depends(get_db)):
    """Valide officiellement un emploi du temps et dé-valide les précédents."""
    # Optionnel : On peut décider de n'avoir qu'un seul EDT validé à la fois
    db.query(models.TimetableResult).update({models.TimetableResult.is_validated: False})
    
    res = db.query(models.TimetableResult).filter(models.TimetableResult.id == res_id).first()
    if not res:
        raise HTTPException(status_code=404, detail="Résultat introuvable")
    
    res.is_validated = True
    db.commit(); db.refresh(res)
    return res

@app.delete("/timetable-results/{res_id}", status_code=204)
def delete_timetable_result(res_id: int, db: Session = Depends(get_db)):
    """Supprime un ancien résultat de l'historique."""
    res = db.query(models.TimetableResult).filter(models.TimetableResult.id == res_id).first()
    if not res:
        raise HTTPException(status_code=404, detail="Résultat introuvable")
    db.delete(res); db.commit()

@app.get("/audit/section/{section_id}")
def audit_section(section_id: int, mode: str = "interactive", db: Session = Depends(get_db)):
    """
    Analyse la qualité de l'emploi du temps pour une section.
    - Si mode == 'interactive', on lit la table de travail SQL (assignments).
    - Sinon, on lit le dernier résultat de l'IA (ALNS, RL...) dans timetable_results.
    """
    try:
        if mode == "interactive":
            # ON AUDITE LA TABLE DE TRAVAIL (Celle modifiée par le Drag & Drop)
            # On récupère toutes les affectations de la DB
            all_db_assignments = db.query(models.Assignment).all()
            
            # On convertit en format compatible avec la logique d'audit
            all_assignments = []
            for a in all_db_assignments:
                all_assignments.append({
                    "id": a.id,
                    "module_part_id": a.module_part_id,
                    "teacher_id": a.teacher_id,
                    "room_id": a.room_id,
                    "slot_id": a.slot_id,
                    "section_id": a.section_id,
                    "td_groups": [{"id": g.id} for g in a.td_groups]
                })
        else:
            # ON AUDITE LA SOLUTION ORIGINALE DE L'IA
            result = db.query(models.TimetableResult)\
                       .filter(models.TimetableResult.algo_type == mode)\
                       .order_by(models.TimetableResult.id.desc())\
                       .first()
            
            if not result:
                return {"section": "N/A", "score": 0, "status": "Aucun résultat IA en base"}
                
            all_assignments = result.data

            
        timeslots = {t.id: t for t in db.query(models.Timeslot).all()}
        module_parts = {mp.id: mp for mp in db.query(models.ModulePart).all()}
        section = db.query(models.Section).filter(models.Section.id == section_id).first()
        if not section: return {"section": f"ID {section_id}", "score": 0, "status": "Section non trouvée en base"}
        
        td_groups = db.query(models.TDGroup).filter(models.TDGroup.section_id == section_id).all()
        td_group_ids = [g.id for g in td_groups]
        
        cm_assigns = []
        group_assigns = {gid: [] for gid in td_group_ids}
        
        count_found = 0
        for a in all_assignments:
            mp_id = a.get('module_part_id')
            mp_obj = module_parts.get(mp_id)
            is_really_cm = mp_obj.type.upper() == "CM" if mp_obj else False
            
            a_sid = int(a.get('section_id', -1))
            # Robustesse : g peut être un dict {"id":...} ou un int directement
            a_groups = [int(g['id'] if isinstance(g, dict) else g) for g in a.get('td_groups', [])]
            matched_groups = [gid for gid in td_group_ids if gid in a_groups]
            
            if is_really_cm and a_sid == section_id:
                cm_assigns.append(a)
                count_found += 1
            elif matched_groups: 
                for gid in matched_groups:
                    group_assigns[gid].append(a)
                    count_found += 1

        if count_found == 0:
            return {"section": section.name, "score": 0, "status": "Aucune séance trouvée dans le JSON"}

        # Indices de satisfaction par groupe
        group_details = {
            "compacite": [],
            "pause_dejeuner": [],
            "rythme_fatigue": [],
            "pedagogie_cm": [],
            "path_late_hits_raw": []
        }
        group_satisfaction = []
        
        # Ordre fixe des débuts de cours pour détecter les trous
        time_order = ["08:30", "10:35", "12:30", "14:30", "16:35"]
        
        for gid in td_group_ids:
            my_path = cm_assigns + group_assigns[gid]
            if not my_path: continue
            
            path_gaps = 0
            path_lunch_hits = 0
            path_late_hits = 0
            day_map = {}
            for a in my_path:
                sid = a.get('slot_id')
                if sid is not None:
                    ts = timeslots.get(int(sid))
                    if ts: day_map.setdefault(ts.day.upper(), []).append(ts)
            
            for day, day_slots in day_map.items():
                day_slots.sort(key=lambda x: x.start_time)
                
                # GAPS calculation (Excluding natural 12:30 lunch break)
                if len(day_slots) > 1:
                    indices = []
                    for ts in day_slots:
                        t_str = ts.start_time.strftime("%H:%M")
                        if t_str in time_order:
                            indices.append(time_order.index(t_str))
                    if len(indices) > 1:
                        # Gaps are missing slots between classes
                        # But we ignore index 2 (12:30) as it's a standard lunch break
                        should_be = set(range(min(indices), max(indices) + 1))
                        missing = should_be - set(indices)
                        true_gaps = len([m for m in missing if m != 2])
                        path_gaps += true_gaps
                
                # LUNCH & FATIGUE calculation
                for ts in day_slots:
                    start_str = ts.start_time.strftime("%H:%M")
                    if "12:30" in start_str: path_lunch_hits += 1
                    if "16:35" in start_str: path_late_hits += 1
            
            # --- NOUVELLE LOGIQUE DE CALCUL ---
            
            # 1. Lunch Analysis (Détection de Tunnel)
            day_tunnels = 0
            for day, day_slots in day_map.items():
                start_times = [ts.start_time.strftime("%H:%M") for ts in day_slots]
                # Tunnel : 10:35 ET 12:30 ET 14:30 (Pénalité forte mais un peu moins radicale)
                if all(t in start_times for t in ["10:35", "12:30", "14:30"]):
                    day_tunnels += 1
            
            # Réduction des pénalités (10 pts par cours à midi, 20 pts par tunnel)
            l_score = max(0, 100 - (path_lunch_hits * 10) - (day_tunnels * 20))

            # 2. Gaps (Compactness) - PASSAGE À 10 POINTS DE PÉNALITÉ AU LIEU DE 20
            g_score = max(0, 100 - (path_gaps * 10))

            # 3. Fatigue (Late & Saturday)
            f_score = max(0, 100 - (path_late_hits * 10))
            if "SAMEDI" in [d.upper() for d in day_map.keys()]:
                sat_slots = day_map.get("SAMEDI", day_map.get("Samedi", []))
                sat_hours = len(sat_slots)
                f_score = max(0, f_score - (20 + sat_hours * 10))
            
            # 4. Stability (Pedagogy: CM in morning)
            cm_slots = [timeslots.get(int(a['slot_id'])) for a in cm_assigns if a.get('slot_id') is not None]
            total_cms = len(cm_assigns)
            
            total_weighted_stability = 0
            for ts in cm_slots:
                if not ts: continue
                # On normalise l'heure pour éviter les problèmes de secondes
                time_str = ts.start_time.strftime("%H:%M")
                
                if time_str in ["08:30", "10:35"]:
                    total_weighted_stability += 100
                elif time_str in ["12:30", "14:30"]:
                    total_weighted_stability += 60  # Pénalité plus forte (40% de perte)
                else: 
                    # Créneau de 16:35 ou plus tard (Inacceptable pour un CM)
                    total_weighted_stability += 0
            
            s_score = round(total_weighted_stability / total_cms, 1) if total_cms > 0 else 100.0

            group_details["compacite"].append(g_score)
            group_details["pause_dejeuner"].append(l_score)
            group_details["rythme_fatigue"].append(f_score)
            group_details["pedagogie_cm"].append(s_score)
            group_details["path_late_hits_raw"].append(path_late_hits)
            
            # Weighted average: Lunch (35%), Gaps (35%), Fatigue (20%), Stability (10%)
            group_satisfaction.append((g_score * 0.35 + l_score * 0.35 + f_score * 0.2 + s_score * 0.1))

        if not group_satisfaction:
            return {"section": section.name, "score": 0, "status": "Pas de données"}
            
        # Calcul des moyennes par indicateur
        avg_gaps = sum(group_details["compacite"]) / len(group_details["compacite"])
        avg_lunch = sum(group_details["pause_dejeuner"]) / len(group_details["pause_dejeuner"])
        avg_fatigue = sum(group_details["rythme_fatigue"]) / len(group_details["rythme_fatigue"])
        avg_stability = sum(group_details["pedagogie_cm"]) / len(group_details["pedagogie_cm"])
        
        # Le seuil s'applique au score de Pédagogie (Stability)
        total_late_sessions = sum(group_details.get("path_late_hits_raw", [0])) / len(td_group_ids)
        if 30 <= avg_stability < 55 and total_late_sessions <= 3:
            avg_stability = 55.0
            
        # Recalcul du score global pondéré : Lunch (35%), Gaps (35%), Fatigue (20%), Stability (10%)
        avg_score = (avg_gaps * 0.35 + avg_lunch * 0.35 + avg_fatigue * 0.2 + avg_stability * 0.1)

        return {
            "section": section.name,
            "score":  round(avg_score, 1),
            "details": {
                "compacite": round(avg_gaps, 1),
                "pause_dejeuner": round(avg_lunch, 1),
                "rythme_fatigue": round(avg_fatigue, 1),
                "pedagogie_cm": round(avg_stability, 1)
            },
            "status": "Excellent" if avg_score > 85 else "Bon" if avg_score > 70 else "Moyen" if avg_score >= 50 else "Médiocre"
        }
    except Exception as e:
        return {"section": "Erreur", "score": 0, "status": str(e), "details": {}}

@app.post("/assignments/reset")
def reset_assignments(mode: str = None, db: Session = Depends(get_db)):
    """
    Restaure le dernier résultat validé de l'IA dans la table de travail assignments.
    Si 'mode' est fourni (ex: 'ga_sa', 'alns', 'rl'), on restaure ce mode précisément.
    Sinon, on prend le dernier résultat validé (tous modes confondus).
    """
    try:
        query = db.query(models.TimetableResult).filter(models.TimetableResult.is_validated == False)
        
        if mode:
            mode = mode.lower()  # Normalisation : 'GA_SA' -> 'ga_sa'
            res = query.filter(models.TimetableResult.algo_type == mode).order_by(models.TimetableResult.id.desc()).first()
        else:
            res = query.order_by(models.TimetableResult.id.desc()).first()

        if not res:
            raise HTTPException(status_code=404, detail=f"Aucun résultat validé trouvé{' pour le mode ' + mode if mode else ''}.")

        # Réinitialise les affectations actuelles
        all_assignments = db.query(models.Assignment).all()
        for a in all_assignments:
            a.slot_id = None
            a.room_id = None
            a.td_groups = []
        db.flush()

        # Ré-applique les données du résultat IA
        for item in res.data:
            assignment_id = item.get('id')
            if not assignment_id: continue
            db_a = db.query(models.Assignment).filter(models.Assignment.id == assignment_id).first()
            if not db_a: continue
            db_a.room_id = item.get('room_id')
            db_a.slot_id = item.get('slot_id')
            db_a.teacher_id = item.get('teacher_id')
            db_a.is_locked = item.get('is_locked', False)
            db_a.section_id = item.get('section_id')
            for g_item in item.get('td_groups', []):
                g_id = g_item['id'] if isinstance(g_item, dict) else g_item
                g_obj = db.query(models.TDGroup).filter(models.TDGroup.id == g_id).first()
                if g_obj:
                    db_a.td_groups.append(g_obj)

        db.commit()
        return {"message": f"Planning '{res.algo_type}' restauré avec succès."}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur lors de la restauration : {str(e)}")

import subprocess
import os
import sys

@app.post("/run-algorithm/{algo}")
def run_algorithm(algo: str):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    algo = algo.lower()
    
    scripts = {
        "ga_sa": os.path.join(base_dir, "algorithms", "1-ga_sa_hybrid", "v2", "main_solver.py"),
        "alns": os.path.join(base_dir, "algorithms", "2-ILS-ALNS", "main_alns.py"),
        "rl": os.path.join(base_dir, "algorithms", "5-RL-ALNS-Curriculum", "main_fused.py")
    }
    
    script_path = scripts.get(algo)
    if not script_path or not os.path.exists(script_path):
        raise HTTPException(status_code=404, detail=f"Script introuvable pour {algo}. Cherché à: {script_path}")
        
    try:
        # Lance le script dans un terminal séparé (Visible sous Windows !)
        if os.name == 'nt':
            # Sous Windows, CREATE_NEW_CONSOLE va faire poper (ouvrir) une fenêtre CMD indépendante
            CREATE_NEW_CONSOLE = 0x00000010
            subprocess.Popen([sys.executable, script_path], creationflags=CREATE_NEW_CONSOLE)
            return {"message": f"Terminal Python ouvert pour l'algorithme {algo}.", "logs": "Lancement dans un nouveau terminal en cours..."}
        else:
            subprocess.Popen([sys.executable, script_path])
            return {"message": f"Processus {algo} lancé en arrière-plan.", "logs": "Voir les logs du serveur"}
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur de lancement : {str(e)}")

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import time as dt_time

@app.get("/export-excel")
async def export_excel(id: int):
    """Génère et télécharge le fichier Excel pour un résultat spécifique."""
    try:
        # Appel du script externe (python export_excel_db.py <ID>)
        cmd = [sys.executable, "export_excel_db.py", str(id)]
        subprocess.run(cmd, check=True)
        
        file_path = f"export_result_{id}.xlsx"
        if os.path.exists(file_path):
            return FileResponse(file_path, filename=f"Emploi_du_temps_{id}.xlsx")
        else:
            raise HTTPException(status_code=500, detail="Fichier Excel non généré.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/export-tp-excel/{section_id}")
async def export_tp_excel(section_id: int, merge: bool = True, db: Session = Depends(get_db)):
    try:
        # 1. Charger les données essentielles
        section = db.query(models.Section).filter(models.Section.id == section_id).first()
        if not section: raise HTTPException(404, "Section non trouvée")
        
        tp_data = json.load(open(os.path.join(TP_CONFIG_DIR, f"section_{section_id}.json"), "r", encoding="utf-8")) if os.path.exists(os.path.join(TP_CONFIG_DIR, f"section_{section_id}.json")) else []
        tp_config_raw = tp_data.get('assignments', []) if isinstance(tp_data, dict) else tp_data
        tp_config = [a for a in tp_config_raw if a.get('slot_id') is not None]
        module_parts = {p.id: p for p in db.query(models.ModulePart).all()}
        modules = {m.id: m for m in db.query(models.Module).all()}
        timeslots = {t.id: t for t in db.query(models.Timeslot).all()}
        td_groups_dict = {g.id: g for g in db.query(models.TDGroup).all()}

        wb = Workbook()
        ws = wb.active
        ws.title = "Planning TP"
        
        # Styles Premium
        thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Entête Noir Premium
        ws.merge_cells("A1:C1")
        ws["A1"] = f"FSTG Timetabling - EMPLOI DU TEMPS TP - {section.name.upper()}"
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # Heures Colonnes (BLOCS DE 4H)
        ws.cell(row=2, column=2, value="MATIN (08h30 - 12h30)")
        ws.cell(row=2, column=3, value="APRES-MIDI (14h30 - 18h30)")
        
        header_fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
        for c in [2, 3]:
            cell = ws.cell(row=2, column=c)
            cell.fill = header_fill
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border
        ws.row_dimensions[2].height = 30
        
        color_map = {
            "Biologie Animale et Végétale": "D1FAE5", # Vert
            "Géodynamique externe": "DBEAFE",       # Bleu
            "Réactivité Chimique": "F3F4F6",         # Gris
            "Culture digitale": "F3E8FF",           # Mauve
            "Mécanique des Fluides": "FEF3C7"        # Orange
        }

        days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]
        current_row = 3

        for day in days:
            # Cellule Jour
            ws.row_dimensions[current_row].height = 100
            day_cell = ws.cell(row=current_row, column=1, value=day.upper())
            day_cell.font = Font(bold=True, color="FFFFFF", size=12)
            day_cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
            day_cell.alignment = Alignment(vertical="center", horizontal="center")
            day_cell.border = thin_border
            
            # Pour chaque bloc de 4h (Matin / AM)
            for b_idx in [0, 1]:
                col = b_idx + 2
                is_morning = (b_idx == 0)
                
                # Récupérer TOUS les TP du bloc de 4h
                cell_content = []
                # Regrouper par module pour éviter les répétitions si merge=True
                mod_buckets = {}
                for asgn in tp_config:
                    ts = timeslots.get(int(asgn['slot_id']))
                    if not ts or ts.day.lower() != day.lower(): continue
                    
                    h = ts.start_time.hour if isinstance(ts.start_time, dt_time) else int(str(ts.start_time)[:2])
                    match = (h < 13) if is_morning else (h >= 13)
                    
                    if match:
                        mp = module_parts.get(asgn.get('module_part_id'))
                        if mp: 
                            key = mp.module_id if merge else str(asgn.get('id', id(asgn)))
                            mod_buckets.setdefault(key, []).append(asgn)

                # Construire le texte épuré fidèle à l'UI
                for key_id, assigns in mod_buckets.items():
                    mp_id = assigns[0].get('module_part_id')
                    mp = module_parts.get(mp_id)
                    m_obj = modules.get(mp.module_id) if mp else None
                    m_name = m_obj.name if m_obj else "TP"
                    
                    # On veut : "Gr X A & Gr Y A / Gr X B & Gr Y B"
                    assignment_labels = []
                    for a in assigns:
                        gids = [gid.get('id') if isinstance(gid, dict) else gid for gid in a.get('td_groups', [])]
                        sub = a.get('tp_subgroup', "") # A, B, C...
                        gnames = []
                        for gid in gids:
                            g_obj = td_groups_dict.get(gid)
                            if g_obj:
                                # Extraire juste le chiffre "1", "2"...
                                g_num = "".join(filter(str.isdigit, g_obj.name.replace(section.name, "")))
                                gnames.append(f"Gr {g_num}{sub}")
                        
                        assignment_labels.append(" & ".join(gnames))
                    
                    alt_text = "(par alternance)" if len(assigns) > 1 else "(par quinzaine)"
                    alt_text = alt_text if merge else ""
                    
                    cell_content.append(f"● {m_name}\n   {' / '.join(assignment_labels)} {alt_text}".strip())

                cell = ws.cell(row=current_row, column=col, value="\n".join(cell_content))
                cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center", indent=1)
                cell.font = Font(size=10, bold=True)
                cell.border = thin_border
                
                # Couleur de fond discrète si contient des données
                if cell_content:
                    cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            current_row += 1

        # Ajustements finaux
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 60

        export_path = f"export_tp_{section_id}_clean.xlsx"
        wb.save(export_path)
        return FileResponse(export_path, filename=f"Planning_TP_Clean_{section.name}.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, str(e))
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
@app.get("/debug/gb-s4")
def debug_gb_s4(db: Session = Depends(get_db)):
    """Récupère toutes les affectations pour GB S4 Gr 1."""
    # 1. Trouver les IDs liés à GB S4
    sections = db.query(models.Section).filter(models.Section.name.like("%GB S4%")).all()
    sec_ids = [s.id for s in sections]
    
    # 2. Récupérer les assignments CM
    cm_asgns = db.query(models.Assignment).filter(models.Assignment.section_id.in_(sec_ids)).all()
    
    results = []
    for a in cm_asgns:
        mp = db.query(models.ModulePart).get(a.module_part_id)
        mod = db.query(models.Module).get(mp.module_id) if mp else None
        ts = db.query(models.Timeslot).get(a.slot_id) if a.slot_id else None
        results.append({
            "id": a.id, "module": mod.name if mod else "?", "type": mp.type if mp else "?",
            "day": ts.day if ts else "?", "time": ts.start_time if ts else "?",
            "info": "CM"
        })
@app.get("/debug/duplicate-td-to-gr2")
def duplicate_td_to_gr2(db: Session = Depends(get_db)):
    """Duplique les TD du Groupe 1 vers le Groupe 2 pour GB S4."""
    # 1. Identifier GB S4
    section = db.query(models.Section).filter(models.Section.name.like("%GB S4%")).first()
    if not section: return {"error": "Section GB S4 non trouvee"}
    
    # 2. Groupe Cible : ID #35 (Gr 2)
    target_group_id = 35
    
    # 3. Récupérer les TD actuels de la section (qui sont implicitement Gr 1 pour l'instant)
    td_asgns = db.query(models.Assignment).join(models.ModulePart).filter(
        models.Assignment.section_id == section.id,
        models.ModulePart.type == "TD"
    ).all()
    
    new_count = 0
    for a in td_asgns:
        # Créer une copie
        new_a = models.Assignment(
            module_part_id=a.module_part_id,
            teacher_id=a.teacher_id,
            room_id=a.room_id,
            slot_id=a.slot_id,
            section_id=a.section_id,
            is_locked=a.is_locked
        )
        db.add(new_a)
        db.flush() # Pour avoir l'ID
        
        # Lier au groupe 35 dans la table d'association (si elle existe)
        # Note: Dans votre modèle, assignments et td_groups sont liés par une table secondary
        target_group = db.query(models.TDGroup).get(target_group_id)
        if target_group:
            new_a.td_groups.append(target_group)
            new_count += 1
            
    db.commit()
    return {"message": f"Succès : {new_count} séances de TD dupliquées pour le Groupe 2.", "section": section.name}

@app.get("/emergency-migrate")
def emergency_migrate(db: Session = Depends(get_db)):
    """Route de secours pour ajouter la colonne manquante si le schéma est vieux."""
    from sqlalchemy import text
    try:
        db.execute(text("ALTER TABLE timetable_results ADD COLUMN is_master_reference BOOLEAN DEFAULT FALSE;"))
        db.commit()
        return {"message": "Success: Column is_master_reference added."}
    except Exception as e:
        db.rollback()
        return {"message": f"Notice/Error: {str(e)}"}
