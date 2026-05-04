import json
import os
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API = "http://localhost:8000"
BASE_DIR = r"c:\Users\HP\OneDrive\Bureau\pfe\_Project_PFE"
CACHE_PATH = os.path.join(BASE_DIR, "data_cache.json")
TIMETABLE_PATH = os.path.join(BASE_DIR, "backend", "generated_timetable.json")
OUTPUT_PATH = os.path.join(BASE_DIR, "algorithms", "ga_sa_hybrid", "v2", "logs", "FSTG_EXCEL_PREMIUM.xlsx")

def set_border(ws, r1, r2, c1, c2, style='thin'):
    s = Side(style=style, color="333333")
    for r in range(r1, r2+1):
        for c in range(c1, c2+1): ws.cell(r, c).border = Border(left=s, right=s, top=s, bottom=s)

def norm_day(d):
    d = str(d).upper().strip()
    return {"MONDAY":"LUNDI","TUESDAY":"MARDI","WEDNESDAY":"MERCREDI","THURSDAY":"JEUDI","FRIDAY":"VENDREDI","SATURDAY":"SAMEDI"}.get(d,d)

print("--- GENERATION EXCEL PREMIUM : Vue Globale + Sections ---")

with open(CACHE_PATH, "r", encoding="utf-8") as f: cache = json.load(f)
with open(TIMETABLE_PATH, "r", encoding="utf-8") as f: timetable = json.load(f)

# --- 1. MAPPINGS ---
group_to_sid = {str(g['id']): str(g['section_id']) for g in cache.get('td-groups', [])}
rooms_list = sorted(cache.get('rooms', []), key=lambda x: (0 if x['type']=='AMPHI' else 1, x['name']))
rooms_dict = {str(r['id']): r for r in rooms_list}
slots_lookup = {str(s['id']): s for s in cache.get('timeslots', [])}
sections_dict = {str(s['id']): s['name'] for s in cache.get('sections', [])}
mp_dict = {str(m['id']): m for m in cache.get('module-parts', [])}
try: mod_names = {str(m['id']): m['name'] for m in requests.get(f"{API}/modules", timeout=5).json()}
except: mod_names = {}

module_audience = {}
for entry in timetable:
    mid = str(mp_dict.get(str(entry.get('module_part_id')),{}).get('module_id'))
    if mid == 'None': continue
    sid = str(entry.get('section_id'))
    if sid!='None': module_audience.setdefault(mid, set()).add(sid)
    for g in entry.get('td_groups', []):
        gid = str(g.get('id'))
        if gid in group_to_sid: module_audience.setdefault(mid, set()).add(group_to_sid[gid])

FSTG_LABELS = ["08h30 - 10h25", "10h35 - 12h30", "12h30 - 14h25", "14h30 - 16h25", "16h35 - 18h30"]
TIME_MAP = {"08:30:00":0, "10:35:00":1, "12:30:00":2, "14:30:00":3, "16:35:00":4}
ALL_DAYS = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI"]
COLORS = ["D6EAF8", "D5F5E3", "FCF3CF", "FADBD8", "EBDEF0", "F9E79F", "F5CBA7", "D1F2EB", "AED6F1", "A2D9CE"]

wb = Workbook(); wb.remove(wb.active)

# --- 2. VUE GLOBALE SALLES ---
ws_g = wb.create_sheet("VUE_GLOBALE_SALLES")
ws_g.sheet_view.showGridLines = False
ws_g.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rooms_list)+1)
ws_g.row_dimensions[1].height = 40
t = ws_g.cell(1, 1, "   PLANNING GLOBAL DES SALLES (FSTG)")
t.font=Font(bold=True, size=16, color="FFFFFF"); t.fill=PatternFill("solid", start_color="1B2631"); t.alignment=Alignment(vertical="center")

for i, r in enumerate(rooms_list, 2):
    c = ws_g.cell(2, i, r['name'])
    c.font=Font(bold=True); c.alignment=Alignment(horizontal='center'); ws_g.column_dimensions[get_column_letter(i)].width = 22

curr = 3
for dname in ALL_DAYS:
    ws_g.merge_cells(start_row=curr, start_column=1, end_row=curr, end_column=len(rooms_list)+1)
    dj = ws_g.cell(curr, 1, dname)
    dj.font=Font(bold=True, color="FFFFFF"); dj.fill=PatternFill("solid", "2E4057"); dj.alignment=Alignment(horizontal='center')
    curr += 1
    for h_str in ["08:30:00", "10:35:00", "12:30:00", "14:30:00", "16:35:00"]:
        ws_g.cell(curr, 1, h_str[:5]).font=Font(bold=True); ws_g.cell(curr, 1).alignment=Alignment(vertical='center')
        for entry in timetable:
            sl = slots_lookup.get(str(entry.get('slot_id')), {})
            if norm_day(sl.get('day',''))==dname and sl.get('start_time','')==h_str:
                rid = str(entry.get('room_id'))
                if rid in rooms_dict:
                    col_idx = [rm['id'] for rm in rooms_list].index(int(rid)) + 2
                    mid = str(mp_dict.get(str(entry.get('module_part_id',{})), {}).get('module_id'))
                    sname = sections_dict.get(str(entry.get('section_id')), "Section")
                    cell = ws_g.cell(curr, col_idx, f"{mod_names.get(mid,'Module')}\n({sname})")
                    cell.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True); cell.font=Font(size=9)
                    cell.fill=PatternFill("solid", COLORS[col_idx % len(COLORS)]); set_border(ws_g, curr, curr, col_idx, col_idx)
        ws_g.row_dimensions[curr].height = 50
        curr += 1

# --- 3. SECTIONS ---
section_results = {sid: {} for sid in sections_dict}
for entry in timetable:
    sl = slots_lookup.get(str(entry.get('slot_id')), {})
    d, t_idx = norm_day(sl.get('day','')), TIME_MAP.get(str(sl.get('start_time','')), -1)
    if d not in ALL_DAYS or t_idx == -1: continue
    mp = mp_dict.get(str(entry.get('module_part_id')), {})
    mid, m_type = str(mp.get('module_id')), mp.get('type','CM') or 'CM'
    m_name = mod_names.get(mid, "Module"); r_name = rooms_dict.get(str(entry.get('room_id')), {}).get('name','Salle')
    targets = set()
    if m_type == 'CM': targets.update(module_audience.get(mid, set()))
    else:
        if str(entry.get('section_id')) != 'None': targets.add(str(entry.get('section_id')))
        for g in entry.get('td_groups', []):
            gid = str(g.get('id')); 
            if gid in group_to_sid: targets.add(group_to_sid[gid])
    for sid in targets:
        if sid in section_results:
            existing = section_results[sid].get((d, FSTG_LABELS[t_idx]), [])
            if not any(x['m'] == m_name and x['t'] == m_type for x in existing):
                section_results[sid].setdefault((d, FSTG_LABELS[t_idx]), []).append({"m": m_name, "t": m_type, "r": r_name})

for sid, name in sections_dict.items():
    ws = wb.create_sheet(name[:30].replace("/","_"))
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8); ws.row_dimensions[1].height = 40
    h = ws.cell(1, 1, f"   FSTG - EMPLOI DU TEMPS - {name.upper()}")
    h.font=Font(bold=True,size=15,color="FFFFFF"); h.fill=PatternFill("solid","1B2631"); h.alignment=Alignment(vertical="center")
    for i, lbl in enumerate(FSTG_LABELS):
        col = {0:2, 1:3, 2:5, 3:7, 4:8}[i]
        c = ws.cell(3, col, lbl); c.font=Font(bold=True); c.alignment=Alignment(horizontal='center'); c.fill=PatternFill("solid","F9E79F"); set_border(ws, 3,3, col, col)
    data = section_results.get(sid, {})
    for di, dname in enumerate(ALL_DAYS):
        rs = 4 + (di * 12)
        ws.merge_cells(start_row=rs, start_column=1, end_row=rs+11, end_column=1)
        dj=ws.cell(rs,1,dname); dj.font=Font(bold=True); dj.alignment=Alignment(horizontal='center',vertical='center',text_rotation=90); set_border(ws,rs,rs+11,1,1)
        for i in range(5):
            col = {0:2, 1:3, 2:5, 3:7, 4:8}[i]
            assigns = data.get((dname, FSTG_LABELS[i]), [])
            if not assigns: set_border(ws, rs, rs+11, col, col)
            else:
                bk = 12 // len(assigns)
                for ai, a in enumerate(assigns):
                    r1, r2 = rs+ai*bk, rs+(ai+1)*bk-1 if ai<len(assigns)-1 else rs+11
                    ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)
                    cell=ws.cell(r1, col, f"{a['m']}\n({a['t']} - {a['r']})")
                    cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cell.font=Font(bold=True,size=11); cell.fill=PatternFill("solid",COLORS[abs(hash(a['m']))%len(COLORS)]); set_border(ws, r1, r2, col, col)
        for r in range(rs, rs+12):
            ws.row_dimensions[r].height = 15
            for sc in [4, 6]: ws.cell(r, sc).fill = PatternFill("solid", start_color="D5DBDB"); set_border(ws, r, r, sc, sc)
    for c in [2,3,5,7,8]: ws.column_dimensions[get_column_letter(c)].width = 30
    ws.column_dimensions['A'].width = 8; ws.column_dimensions['D'].width = 0.5; ws.column_dimensions['F'].width = 0.5

wb.save(OUTPUT_PATH)
print("✅ EXCEL PREMIUM GENERE AVEC SUCCES !")
