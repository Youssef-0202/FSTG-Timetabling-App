import json
import os
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API = "http://localhost:8000"
# Utilisation de chemins dynamiques basés sur l'emplacement du script
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CACHE_PATH = os.path.join(BASE_DIR, "data_cache.json")
TIMETABLE_PATH = os.path.join(BASE_DIR, "backend", "generated_timetable_rl.json")
OUTPUT_PATH = os.path.join(BASE_DIR, "algorithms", "rl_controller", "logs", "FSTG_EXCEL_RL_PREMIUM.xlsx")

FSTG_LABELS = ["08h30 - 10h25", "10h35 - 12h30", "12h30 - 14h25", "14h30 - 16h25", "16h35 - 18h30"]
TIME_MAP = {"08:30:00": 0, "10:35:00": 1, "12:30:00": 2, "14:30:00": 3, "16:35:00": 4}
ALL_DAYS = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI"]
SUB_ROWS = 12

COLORS = ["EBF5FB", "E8F8F5", "FEF9E7", "FDF2E9", "F4ECF7", "EBEDEF", "D6EAF8", "D1F2EB", "FCF3CF", "F5CBA7"]

def set_border(ws, r1, r2, c1, c2, style='thin'):
    side = Side(style=style, color="333333")
    border = Border(left=side, right=side, top=side, bottom=side)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1): ws.cell(r, c).border = border

def norm_day(d):
    d = str(d).upper().strip()
    return {"MONDAY": "LUNDI", "TUESDAY": "MARDI", "WEDNESDAY": "MERCREDI",
            "THURSDAY": "JEUDI", "FRIDAY": "VENDREDI", "SATURDAY": "SAMEDI"}.get(d, d)

def run_export():
    if not os.path.exists(TIMETABLE_PATH):
        print(f"ERREUR : Le fichier RL {TIMETABLE_PATH} n'existe pas encore. Lancez main_rl.py d'abord.")
        return

    with open(CACHE_PATH, "r", encoding="utf-8") as f: cache = json.load(f)
    with open(TIMETABLE_PATH, "r", encoding="utf-8") as f: timetable = json.load(f)

    sections_dict = {str(s['id']): s['name'] for s in cache.get('sections', [])}
    rooms_dict = {str(r['id']): r['name'] for r in cache.get('rooms', [])}
    slots_lookup = {str(s['id']): s for s in cache.get('timeslots', [])}
    groups_list = cache.get('td-groups', [])
    group_to_sid = {str(g['id']): str(g['section_id']) for g in groups_list}
    mp_lookup = {str(m['id']): m for m in cache.get('module-parts', [])}
    
    try: 
        mod_names = {str(m['id']): m['name'] for m in requests.get(f"{API}/modules", timeout=5).json()}
    except: 
        print("WAWRNING : API Inaccessible, utilisation des IDs.")
        mod_names = {}

    module_colors = {}
    def get_mod_color(mid):
        if mid not in module_colors: module_colors[mid] = COLORS[len(module_colors)%len(COLORS)]
        return module_colors[mid]

    # --- EXTRACTION ---
    section_data = {sid: {} for sid in sections_dict}
    for entry in timetable:
        sl = slots_lookup.get(str(entry.get('slot_id')), {})
        day, time = norm_day(sl.get('day','')), sl.get('start_time','')
        if day not in ALL_DAYS or time not in TIME_MAP: continue
        mp = mp_lookup.get(str(entry.get('module_part_id')))
        if not mp: continue
        mid_str = str(mp.get('module_id'))
        m_name, m_type = mod_names.get(mid_str, f"M {mid_str}"), mp.get('type', 'CM') or 'CM'
        r_name = rooms_dict.get(str(entry.get('room_id')), "Salle")
        slot_key = (day, FSTG_LABELS[TIME_MAP[time]])
        
        target_sids = set()
        if entry.get('section_id') and str(entry.get('section_id')) != 'None': 
            target_sids.add(str(entry.get('section_id')))
        
        for g in entry.get('td_groups', []) or []:
            gid = str(g.get('id') if isinstance(g, dict) else g)
            if gid in group_to_sid: target_sids.add(group_to_sid[gid])
            
        for sid in target_sids:
            gnames = []
            if m_type != 'CM':
                for g in entry.get('td_groups', []) or []:
                    gid = str(g.get('id') if isinstance(g, dict) else g)
                    if group_to_sid.get(gid) == sid:
                        gnames.append(next((gx for gx in groups_list if str(gx['id'])==gid), {}).get('name','').split(' ')[-1])
            rec = {"m": m_name, "t": m_type, "r": r_name, "grps": gnames, "sid": sid, "mid": mid_str}
            section_data[sid].setdefault(slot_key, []).append(rec)

    wb = Workbook(); wb.remove(wb.active)

    def apply_header(ws, title):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws.row_dimensions[1].height = 45
        c = ws.cell(1, 1, f"RL-AGENT Timetabling - {title.upper()}")
        c.font=Font(bold=True, size=18, color="FFFFFF"); c.fill=PatternFill("solid","4B1C71") # Violet pour différencier RL
        c.alignment=Alignment(horizontal='center', vertical='center')

    def get_rich_text(rec):
        line1 = rec['m']
        line2 = f"({rec['t']} - {rec['r']})"
        sname = sections_dict.get(rec['sid'], 'N/A')
        if rec['t'] == 'CM': line3 = f"Section: {sname}"
        else: line3 = f"{sname} - Groupe: {', '.join(rec['grps'])}" if rec['grps'] else f"Section: {sname}"
        return f"{line1}\n{line2}\n{line3}".strip()

    # --- 1. VUE GLOBALE ---
    ws_g = wb.create_sheet("VUE_GLOBALE_RL")
    ws_g.sheet_view.showGridLines = False; apply_header(ws_g, "Planning RL (Toutes Salles)")
    sorted_rooms = sorted(cache.get('rooms', []), key=lambda x: (0 if x['type']=='AMPHI' else 1, x['name']))
    for i,r in enumerate(sorted_rooms, 2):
        ws_g.cell(2, i, r['name']).font=Font(bold=True); ws_g.cell(2,i).alignment=Alignment(horizontal='center'); ws_g.column_dimensions[get_column_letter(i)].width=28
    curr = 3
    for dname in ALL_DAYS:
        ws_g.merge_cells(start_row=curr, start_column=1, end_row=curr, end_column=len(sorted_rooms)+1)
        ws_g.cell(curr,1,dname).font=Font(bold=True,color="FFFFFF"); ws_g.cell(curr,1).fill=PatternFill("solid","34495E"); ws_g.cell(curr,1).alignment=Alignment(horizontal='center'); curr+=1
        for h_str in ["08:30:00", "10:35:00", "12:30:00", "14:30:00", "16:35:00"]:
            ws_g.cell(curr,1,h_str[:5]).font=Font(bold=True); ws_g.cell(curr,1).alignment=Alignment(vertical='center')
            for col in range(2, len(sorted_rooms) + 2): set_border(ws_g, curr, curr, col, col)
            for e in timetable:
                sl = slots_lookup.get(str(e.get('slot_id')), {})
                if norm_day(sl.get('day',''))==dname and sl.get('start_time','')==h_str:
                    rid = str(e.get('room_id'))
                    col_idx = next((idx for idx,rm in enumerate(sorted_rooms,2) if str(rm['id'])==rid), None)
                    if col_idx:
                        mp = mp_lookup.get(str(e.get('module_part_id')), {}); m_type = mp.get('type','CM') or 'CM'
                        mid = str(mp.get('module_id')); sid = str(e.get('section_id'))
                        gs = [str(g.get('id') if isinstance(g, dict) else g) for g in e.get('td_groups', []) or []]
                        if sid == 'None' and gs: sid = group_to_sid.get(gs[0], 'None')
                        gnames = [next((gx for gx in groups_list if str(gx['id'])==gid), {}).get('name','').split(' ')[-1] for gid in gs]
                        c = ws_g.cell(curr, col_idx, get_rich_text({"m": mod_names.get(mid,'Mod'), "t": m_type, "r": rooms_dict.get(rid,''), "grps": gnames, "sid": sid}))
                        c.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True); c.font=Font(size=9)
                        c.fill=PatternFill("solid", get_mod_color(mid)); set_border(ws_g, curr, curr, col_idx, col_idx)
            ws_g.row_dimensions[curr].height=70; curr+=1

    # --- 2. VUES SECTIONS ---
    for sid in sorted(section_data.keys(), key=lambda x: sections_dict.get(x, '')):
        name = sections_dict[sid]
        ws = wb.create_sheet(name[:30].replace("/","_"))
        ws.sheet_view.showGridLines = False; apply_header(ws, f"Emploi RL - {name}")
        col_map = {0:2, 1:3, 2:5, 3:7, 4:8}
        for i,l in enumerate(FSTG_LABELS):
            c=ws.cell(3,col_map[i],l); c.font=Font(bold=True); c.alignment=Alignment(horizontal='center'); c.fill=PatternFill("solid","D2B4DE"); set_border(ws,3,3,col_map[i],col_map[i]) # Violet clair
        for di, dn in enumerate(ALL_DAYS):
            rs=4+(di*SUB_ROWS); ws.merge_cells(start_row=rs,start_column=1,end_row=rs+SUB_ROWS-1,end_column=1)
            dj=ws.cell(rs,1,dn); dj.font=Font(bold=True,color="FFFFFF"); dj.fill=PatternFill("solid","2E4053"); dj.alignment=Alignment(horizontal='center',vertical='center',text_rotation=90); set_border(ws,rs,rs+SUB_ROWS-1,1,1)
            for i in range(5):
                col=col_map[i]; assigns=section_data[sid].get((dn,FSTG_LABELS[i]), [])
                if not assigns:
                    ws.merge_cells(start_row=rs, start_column=col, end_row=rs + SUB_ROWS - 1, end_column=col)
                    ws.cell(rs, col).fill = PatternFill("solid", "F2F4F4"); set_border(ws, rs, rs + SUB_ROWS - 1, col, col)
                else:
                    bk=SUB_ROWS//len(assigns); n=len(assigns)
                    for ai,a in enumerate(assigns):
                        r1, r2 = rs+ai*bk, rs+(ai+1)*bk-1 if ai<n-1 else rs+SUB_ROWS-1
                        ws.merge_cells(start_row=r1,start_column=col,end_row=r2,end_column=col)
                        c=ws.cell(r1,col,get_rich_text(a))
                        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.font=Font(bold=True,size=10)
                        c.fill=PatternFill("solid", get_mod_color(a['mid'])); set_border(ws,r1,r2,col,col)
            for sc in [4, 6]:
                ws.merge_cells(start_row=rs, start_column=sc, end_row=rs + SUB_ROWS - 1, end_column=sc)
                ws.cell(rs, sc).fill = PatternFill("solid", "EBEDEF"); set_border(ws, rs, rs + SUB_ROWS - 1, sc, sc)
        for c in [2,3,5,7,8]: ws.column_dimensions[get_column_letter(c)].width=33
        ws.column_dimensions['A'].width=8; ws.column_dimensions['D'].width=0.5; ws.column_dimensions['F'].width=0.5
    
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"DONE: EXCEL RL FINAL - {OUTPUT_PATH}")

if __name__ == "__main__":
    run_export()
